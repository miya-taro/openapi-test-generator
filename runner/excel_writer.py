from __future__ import annotations
from datetime import datetime, timezone
from pathlib import Path
import openpyxl
from models import RunResult

RESULT_HEADERS = [
    "実行日時",        # P
    "実測ステータス",  # Q
    "実測ヘッダ",      # R
    "実測ボディ",      # S
    "応答時間(ms)",    # T
    "判定詳細",        # U
    "実行コマンド",    # V
]

BODY_MAX_LEN = 21


def write_results(
    source_path: str | Path,
    results: list[tuple[RunResult, str, str]],  # (result, verdict, detail)
    output_path: str | Path,
    full_body: bool = False,
) -> None:
    wb = openpyxl.load_workbook(str(source_path))
    ws = wb["test_spec"]

    header_row_idx = 4  # 1-based row index for headers

    # Find the first empty column after existing headers
    header_row = ws[header_row_idx]
    last_col = 0
    for cell in header_row:
        if cell.value is not None:
            last_col = cell.column
    start_col = last_col + 1 if last_col > 0 else 1

    # Write result column headers
    for i, h in enumerate(RESULT_HEADERS):
        ws.cell(row=header_row_idx, column=start_col + i, value=h)

    # Build id → row index map (data starts at row 5)
    id_col = _find_column(ws, header_row_idx, "No")
    id_to_row: dict[str, int] = {}
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=id_col).value
        if cell_val is not None:
            id_to_row[str(cell_val)] = row_idx

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    for result, verdict, detail in results:
        row_idx = id_to_row.get(result.case.id)
        if row_idx is None:
            continue
        body_preview = result.actual_body if full_body else _truncate(result.actual_body, BODY_MAX_LEN)
        ws.cell(row=row_idx, column=start_col + 0, value=now)
        ws.cell(row=row_idx, column=start_col + 1, value=result.actual_status or "")
        ws.cell(row=row_idx, column=start_col + 2, value=_format_headers(result.actual_headers))
        ws.cell(row=row_idx, column=start_col + 3, value=body_preview)
        ws.cell(row=row_idx, column=start_col + 4, value=round(result.elapsed_ms, 1))
        ws.cell(row=row_idx, column=start_col + 5, value=detail)
        ws.cell(row=row_idx, column=start_col + 6, value=result.curl_command)

    wb.save(str(output_path))


def _find_column(ws, row_idx: int, header: str) -> int:
    for cell in ws[row_idx]:
        if cell.value == header:
            return cell.column
    return 1


def _truncate(s: str, max_len: int) -> str:
    if len(s) <= max_len:
        return s
    return s[:max_len] + "…"


def _format_headers(headers: dict[str, str]) -> str:
    return "\n".join(f"{k}: {v}" for k, v in headers.items())
