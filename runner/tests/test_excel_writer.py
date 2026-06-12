import sys
import shutil
import tempfile
from pathlib import Path
import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from excel_writer import write_results
from models import RunCase, RunResult

FIXTURE = Path(__file__).parent / "fixtures" / "sample.xlsx"


def make_case(id: str = "TC-001-001") -> RunCase:
    return RunCase(
        id=id, operation_id="getUsers", path="/users", method="GET",
        param_in="query", param_name="page", input_value="1",
        expected_status="200", expected_response_body="",
        expected_response_header="", expected_response_time="3秒以内",
        expected_result="",
    )


def make_result(case: RunCase, status: int = 200, elapsed_ms: float = 150.0) -> RunResult:
    return RunResult(
        case=case,
        actual_status=status,
        actual_body='{"items":[]}',
        actual_headers={"Content-Type": "application/json"},
        curl_command='curl.exe -s -X GET -D - http://localhost:8080/users?page=1',
        elapsed_ms=elapsed_ms,
        curl_exit_code=0,
        error=None,
    )


@pytest.fixture
def output_file(tmp_path):
    dest = tmp_path / "result.xlsx"
    shutil.copy(str(FIXTURE), str(dest))
    return dest


def test_result_columns_written(output_file):
    case = make_case("TC-001-001")
    result = make_result(case)
    write_results(FIXTURE, [(result, "OK", "")], str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    # Find the 実行日時 header
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    assert "実行日時" in header_row
    assert "実測ステータス" in header_row
    assert "実測ボディ" in header_row
    assert "判定詳細" in header_row
    assert "実行コマンド" in header_row


def test_result_values_in_correct_row(output_file):
    case = make_case("TC-001-001")
    result = make_result(case, status=200, elapsed_ms=150.0)
    write_results(FIXTURE, [(result, "OK", "")], str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]

    # Find column index of 実測ステータス
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    status_col = list(header_row).index("実測ステータス") + 1

    # Data should be in row 5 (first data row)
    val = ws.cell(row=5, column=status_col).value
    assert val == 200


def test_body_truncated_over_21_chars(output_file):
    case = make_case("TC-001-001")
    long_body = "x" * 30
    result = RunResult(
        case=case, actual_status=200, actual_body=long_body,
        actual_headers={}, elapsed_ms=100.0, curl_exit_code=0, error=None,
    )
    write_results(FIXTURE, [(result, "OK", "")], str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    body_col = list(header_row).index("実測ボディ") + 1
    val = str(ws.cell(row=5, column=body_col).value)
    assert len(val) <= 22  # 21 chars + ellipsis


def test_detail_written_for_ng(output_file):
    case = make_case("TC-001-001")
    result = make_result(case, status=500)
    write_results(FIXTURE, [(result, "NG", "ステータス不一致: 期待=200, 実測=500")], str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    detail_col = list(header_row).index("判定詳細") + 1
    val = ws.cell(row=5, column=detail_col).value
    assert "不一致" in str(val)


def test_curl_command_written(output_file):
    case = make_case("TC-001-001")
    result = make_result(case)
    write_results(FIXTURE, [(result, "OK", "")], str(output_file))

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    cmd_col = list(header_row).index("実行コマンド") + 1
    val = ws.cell(row=5, column=cmd_col).value
    assert val == 'curl.exe -s -X GET -D - http://localhost:8080/users?page=1'


def test_full_body_not_truncated(output_file):
    case = make_case("TC-001-001")
    long_body = "x" * 100
    result = RunResult(
        case=case, actual_status=200, actual_body=long_body,
        actual_headers={}, elapsed_ms=100.0, curl_exit_code=0,
        curl_command="", error=None,
    )
    write_results(FIXTURE, [(result, "OK", "")], str(output_file), full_body=True)

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    body_col = list(header_row).index("実測ボディ") + 1
    val = str(ws.cell(row=5, column=body_col).value)
    assert val == long_body


def test_body_truncated_by_default(output_file):
    case = make_case("TC-001-001")
    long_body = "x" * 100
    result = RunResult(
        case=case, actual_status=200, actual_body=long_body,
        actual_headers={}, elapsed_ms=100.0, curl_exit_code=0,
        curl_command="", error=None,
    )
    write_results(FIXTURE, [(result, "OK", "")], str(output_file), full_body=False)

    wb = openpyxl.load_workbook(str(output_file))
    ws = wb["test_spec"]
    header_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    body_col = list(header_row).index("実測ボディ") + 1
    val = str(ws.cell(row=5, column=body_col).value)
    assert len(val) <= 22  # 21 chars + ellipsis
